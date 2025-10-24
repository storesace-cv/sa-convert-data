import json
import importlib
from pathlib import Path

from openpyxl import load_workbook

from app.backend.cardex_schema import CARDEX_FIELD_ORDER
from app.backend.export_validation import MONETARY_FIELDS, validate_export_rows


def test_validate_export_rows_rounds_and_defaults():
    row = {
        "nome": "  Produto   ",
        "preco_compra_fornecedor": "1.005",
        "pvp1_s_com_iva": "2,004",
        "categoria": "",
        "estado": None,
        "extra": "ignored",
    }
    cleaned_rows, metadata = validate_export_rows([row])

    assert metadata.total_rows == 1
    assert metadata.missing_columns["cod_artigo"] == 1
    assert metadata.defaults_applied == [
        {"row": 0, "field": "categoria", "value": "Default"},
        {"row": 0, "field": "estado", "value": "ATIVO"},
    ]
    assert {
        (entry["field"], entry["output"])
        for entry in metadata.rounding_adjustments
    } == {
        ("preco_compra_fornecedor", 1.01),
        ("pvp1_s_com_iva", 2.0),
    }
    assert metadata.extra_fields == [{"row": 0, "fields": ["extra"]}]

    cleaned = cleaned_rows[0]
    assert cleaned["nome"] == "Produto"
    assert cleaned["categoria"] == "Default"
    assert cleaned["estado"] == "ATIVO"
    assert cleaned["preco_compra_fornecedor"] == 1.01
    assert cleaned["pvp1_s_com_iva"] == 2.0
    for field in CARDEX_FIELD_ORDER:
        assert field in cleaned


def test_validate_export_rows_flags_invalid_monetary():
    row = {field: "invalid" for field in MONETARY_FIELDS}
    cleaned_rows, metadata = validate_export_rows([row])

    assert metadata.total_rows == 1
    assert len(metadata.invalid_monetary) == len(MONETARY_FIELDS)
    assert all(entry["row"] == 0 for entry in metadata.invalid_monetary)

    cleaned = cleaned_rows[0]
    for field in MONETARY_FIELDS:
        assert cleaned[field] is None


def test_run_export_validation_cli_generates_artifacts(tmp_path, monkeypatch):
    batch_id = "batch-cli"
    exports_dir = tmp_path / "exports"
    db_path = tmp_path / "data.db"
    db_dir = tmp_path / "dbdir"

    monkeypatch.setenv("SA_CONVERT_DB", str(db_path))
    monkeypatch.setenv("SA_CONVERT_DB_DIR", str(db_dir))
    monkeypatch.setenv("SA_CONVERT_EXPORT_DIR", str(exports_dir))
    monkeypatch.setenv("SA_CONVERT_USER", "cli-user")

    import app.config as config
    import app.backend.db as db_module
    import app.backend.exporter_excel as exporter_module
    import tools.export_validate as cli_module

    config = importlib.reload(config)
    db_module = importlib.reload(db_module)
    exporter_module = importlib.reload(exporter_module)
    cli_module = importlib.reload(cli_module)

    db_module.init_db()
    conn = db_module.connect()
    now = db_module.now_utc()
    with conn:
        cur = conn.execute(
            """
            INSERT INTO imported_raw(
                batch_id, row_index, cod_artigo, cod_barras, nome, nome_norm,
                desc1, desc2, categoria, class_venda,
                unid_default, unid_compra, unid_stock, unid_log,
                estado, created_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                batch_id,
                1,
                "A1",
                "B1",
                "Nome",
                "nome",
                "Desc",
                "Desc",
                "Categoria",
                "X",
                "UN",
                "UN",
                "UN",
                "UN",
                "ATIVO",
                now,
            ),
        )
        raw_id = cur.lastrowid

        cur = conn.execute(
            """
            INSERT INTO canonical_item(name_canonico, scope, rule_version, created_at)
            VALUES (?,?,?,?)
            """,
            ("Canon", "global", "v1", now),
        )
        canonical_id = cur.lastrowid

        cur = conn.execute(
            """
            INSERT INTO cluster_proposal(batch_id, label_sugerido, created_at)
            VALUES (?,?,?)
            """,
            (batch_id, "Sugestao", now),
        )
        cluster_id = cur.lastrowid

        cur = conn.execute(
            """
            INSERT INTO working_article(
                batch_id, raw_id, nome_norm, nome_sem_stop,
                quantidade_valor, quantidade_total, quantidade_unidade,
                quantidade_tipo, quantidade_numero,
                flag_com_sal, flag_sem_sal, marca_detectada
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                batch_id,
                raw_id,
                "nome",
                "nome",
                None,
                None,
                None,
                None,
                None,
                0,
                0,
                None,
            ),
        )
        working_id = cur.lastrowid

        conn.execute(
            """
            INSERT INTO cluster_member(cluster_id, working_id, score, selected_by_user)
            VALUES (?,?,?,?)
            """,
            (cluster_id, working_id, 0.9, 1),
        )

        conn.execute(
            """
            INSERT INTO approval_decision(
                cluster_id, canonical_id, artigo_base_raw_id,
                unit_default, unit_compra, unit_stock, unit_log, decided_at
            ) VALUES (?,?,?,?,?,?,?,?)
            """,
            (
                cluster_id,
                canonical_id,
                raw_id,
                "UN",
                "CX",
                "UN",
                "UN",
                now,
            ),
        )

    conn.close()

    template_path = Path("databases/models/export template.xlsx")
    result = cli_module.run_export_validation(batch_id, model_path=str(template_path))

    excel_path = Path(result["out"])  # reuses exporter path
    report_path = Path(result["report"])

    assert excel_path.exists()
    assert report_path.exists()

    report_data = json.loads(report_path.read_text(encoding="utf-8"))
    assert report_data["batch_id"] == batch_id
    assert report_data["validation"]["total_rows"] == 1

    wb = load_workbook(excel_path)
    try:
        ws = wb.active
        assert ws.max_row >= 3
    finally:
        wb.close()

    conn = db_module.connect()
    try:
        cur = conn.execute(
            "SELECT action, payload_json FROM decision_log WHERE action='export_validation'"
        )
        entry = cur.fetchone()
        assert entry is not None
        payload = json.loads(entry[1])
        assert payload["batch_id"] == batch_id
        assert payload["validation_summary"]["rounding_adjustments"] >= 0
    finally:
        conn.close()
