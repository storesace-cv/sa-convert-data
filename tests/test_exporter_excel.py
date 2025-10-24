import json
from pathlib import Path

from openpyxl import load_workbook

from app.backend.audit import log_action
from app.backend.cardex_schema import CARDEX_FIELD_ORDER
from app.backend.db import connect, init_db, now_utc
from app.backend.exporter_excel import CLASSIFICATION_KEYS, export_excel_using_model


def _insert_sample_dataset(conn, batch_id: str) -> None:
    now = now_utc()
    with conn:
        cur = conn.execute(
            """
            INSERT INTO imported_raw(
                batch_id, row_index, cod_artigo, cod_barras, nome, nome_norm,
                desc1, desc2, desc_curta1, desc_curta2,
                categoria, familia, subfamilia,
                class_venda,
                unid_default, unid_compra, unid_stock, unid_log,
                estado, created_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                batch_id,
                1,
                "123",
                "789",
                "Artigo Base",
                "artigo base",
                "Artigo Base",
                "Artigo Base",
                "Curta 1",
                "Curta 2",
                "Categoria X",
                "Familia Y",
                "Subfamilia Z",
                "X",
                "KG",
                "KG",
                "KG",
                "KG",
                "ATIVO",
                now,
            ),
        )
        raw_id = cur.lastrowid

        cur = conn.execute(
            """
            INSERT INTO canonical_item(name_canonico, scope, rule_version, created_at)
            VALUES (?,?,?,?)
            """,
            ("Produto Canonico", "global", "v1", now),
        )
        canonical_id = cur.lastrowid

        cur = conn.execute(
            """
            INSERT INTO cluster_proposal(batch_id, label_sugerido, created_at)
            VALUES (?,?,?)
            """,
            (batch_id, "Sugestao", now),
        )
        cluster_id = cur.lastrowid

        cur = conn.execute(
            """
            INSERT INTO working_article(
                batch_id, raw_id, nome_norm, nome_sem_stop,
                quantidade_valor, quantidade_total, quantidade_unidade,
                quantidade_tipo, quantidade_numero,
                flag_com_sal, flag_sem_sal, marca_detectada
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                batch_id,
                raw_id,
                "artigo base",
                "artigo base",
                None,
                None,
                None,
                None,
                None,
                0,
                0,
                None,
            ),
        )
        working_id = cur.lastrowid

        conn.execute(
            """
            INSERT INTO cluster_member(cluster_id, working_id, score, selected_by_user)
            VALUES (?,?,?,?)
            """,
            (cluster_id, working_id, 0.9, 0),
        )

        conn.execute(
            """
            INSERT INTO approval_decision(
                cluster_id, canonical_id, artigo_base_raw_id,
                unit_default, unit_compra, unit_stock, unit_log, decided_at
            ) VALUES (?,?,?,?,?,?,?,?)
            """,
            (
                cluster_id,
                canonical_id,
                raw_id,
                "UN",
                "CX",
                "UN",
                "UN",
                now,
            ),
        )


def test_exporter_populates_all_columns_and_log_sheet(tmp_path, monkeypatch):
    db_path = tmp_path / "data.db"
    monkeypatch.setenv("SA_CONVERT_DB", str(db_path))
    monkeypatch.setenv("SA_CONVERT_USER", "ci-user")
    init_db()

    conn = connect()
    batch_id = "batch-123"
    _insert_sample_dataset(conn, batch_id)

    with conn:
        log_action(
            conn,
            f"batch:{batch_id}",
            "cardex_import",
            {
                "batch_id": batch_id,
                "file": "import.xlsx",
                "imported_rows": 1,
                "working_rows": 1,
                "user": "tester",
            },
            ts="2024-01-01T00:00:00Z",
        )

    conn.close()

    template_path = Path("databases/models/export template.xlsx")
    out_path = tmp_path / "export.xlsx"

    result = export_excel_using_model(str(template_path), str(out_path), batch_id)
    assert result["ok"] is True
    assert result["rows"] == 1
    assert Path(result["out"]) == out_path
    assert result["validation"]["total_rows"] == 1
    assert result["cleaned_rows"][0]["nome"] == "Produto Canonico"

    wb = load_workbook(out_path)
    try:
        ws = wb.active
        values = [ws.cell(row=3, column=i).value for i in range(1, len(CARDEX_FIELD_ORDER) + 1)]
        row_map = dict(zip(CARDEX_FIELD_ORDER, values))

        assert row_map["cod_artigo"] is None
        assert row_map["nome"] == "Produto Canonico"
        assert row_map["desc_curta1"] == "Produto Canonico"
        assert row_map["desc_curta2"] == "Produto Canonico"

        assert row_map["class_compra_venda"] == "X"
        for key in CLASSIFICATION_KEYS:
            if key == "class_compra_venda":
                continue
            assert row_map[key] in (None, "")

        assert row_map["unid_default"] == "UN"
        assert row_map["unid_compra"] == "CX"
        assert row_map["unid_stock"] == "UN"
        assert row_map["unid_log"] == "UN"
        assert row_map["categoria"] == "Categoria X"
        assert row_map["estado"] == "ATIVO"

        log_ws = wb["_LOG"]
        meta = {
            log_ws.cell(row=r, column=1).value: log_ws.cell(row=r, column=2).value
            for r in range(2, log_ws.max_row + 1)
        }

        assert meta["source_file"] == "import.xlsx"
        assert meta["import_timestamp"] == "2024-01-01T00:00:00Z"
        assert meta["import_user"] == "tester"
        assert meta["export_user"] == "ci-user"
        assert meta["export_file"] == str(out_path)
        assert meta["clusters_suggested"] == 1
        assert meta["clusters_approved"] == 1
        assert meta["items_excluded"] == 1
        assert meta["validation_rounding_adjustments"] == 0
        assert meta["validation_invalid_monetary"] == 0
        assert meta["validation_defaults_applied"] >= 0
        assert meta["validation_missing_columns"] >= 0

        conn_check = connect()
        try:
            cur = conn_check.execute(
                "SELECT action, payload_json, ts FROM decision_log ORDER BY id"
            )
            entries = cur.fetchall()
            actions = {row[0] for row in entries}
            assert {"cardex_import", "export_generated"}.issubset(actions)
            export_entry = next(
                row
                for row in entries
                if row[0] == "export_generated"
                and json.loads(row[1]).get("batch_id") == batch_id
            )
            payload = json.loads(export_entry[1])
            assert payload["batch_id"] == batch_id
            assert payload["out_file"] == str(out_path)
            assert payload["rows"] == 1
            assert payload["user"] == "ci-user"
            assert meta["export_timestamp"] == export_entry[2]
        finally:
            conn_check.close()
    finally:
        wb.close()
