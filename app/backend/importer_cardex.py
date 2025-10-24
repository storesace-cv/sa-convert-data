import os, glob
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import load_workbook

from app.backend.cardex_schema import (
    CARDEX_FIELD_ORDER,
    CARDEX_IMPORT_COLUMNS,
    imported_raw_columns,
    normalize_header_label,
)
from app.backend.db import connect, init_db, now_utc
from app.backend.text_norm import normalize_name
from app.backend.article_features import compute_article_features
from app.config import IMPORT_DIR, ensure_dirs


def _pick_xlsx(path_or_dir: str) -> str:
  ensure_dirs()
  if not path_or_dir or path_or_dir.strip() == "":
    path_or_dir = str(IMPORT_DIR)
  if os.path.isdir(path_or_dir):
    files = sorted(
      glob.glob(os.path.join(path_or_dir, "*.xlsx")),
      key=os.path.getmtime,
      reverse=True,
    )
    if not files:
      raise FileNotFoundError(f"Nenhum .xlsx encontrado em {path_or_dir}")
    return files[0]
  if not os.path.exists(path_or_dir):
    raise FileNotFoundError(f"Ficheiro não encontrado: {path_or_dir}")
  return path_or_dir


def _resolve_headers(ws) -> tuple[List[Any], int]:
  merged_in_header = any(r.min_row == 1 for r in ws.merged_cells.ranges)
  if merged_in_header:
    super_headers: dict[int, Any] = {}
    for col in range(1, ws.max_column + 1):
      super_headers[col] = ws.cell(row=1, column=col).value
    for rng in ws.merged_cells.ranges:
      if rng.min_row <= 2 and rng.min_row == 1:
        value = ws.cell(row=1, column=rng.min_col).value
        for col in range(rng.min_col, rng.max_col + 1):
          super_headers[col] = value
    headers: List[Any] = []
    for col in range(1, ws.max_column + 1):
      sup = super_headers.get(col)
      sub = ws.cell(row=2, column=col).value
      if sub is not None and str(sub).strip():
        if sup is not None and str(sup).strip() and normalize_header_label(sub) != normalize_header_label(sup):
          headers.append(f"{sup}::{sub}")
        else:
          headers.append(sub)
      else:
        headers.append(sup)
    return headers, 3
  headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
  return headers, 2


def _ensure_column_alignment(headers: Sequence[Any]) -> List[str]:
  if len(headers) < len(CARDEX_IMPORT_COLUMNS):
    raise ValueError("Número de colunas inferior ao esperado para o modelo de importação")
  keys: List[str] = []
  for idx, spec in enumerate(CARDEX_IMPORT_COLUMNS):
    raw = headers[idx]
    norm = normalize_header_label(raw)
    if norm not in spec.normalized_variants():
      raise KeyError(f"Coluna inesperada na posição {idx+1}: {raw!r}")
    keys.append(spec.key)
  return keys


def _row_empty(values: Iterable[Any]) -> bool:
  for v in values:
    if v is None:
      continue
    if isinstance(v, str) and v.strip() == "":
      continue
    return False
  return True


def import_cardex_reformulado(xlsx_path: str | None, batch_id: str) -> Dict[str, Any]:
  init_db()
  xlsx_path = _pick_xlsx(xlsx_path or "")
  wb = load_workbook(filename=xlsx_path, data_only=True)
  try:
    ws = wb.active
    headers, start_row = _resolve_headers(ws)
    column_keys = _ensure_column_alignment(headers)
    insert_columns = imported_raw_columns()
    placeholders = ",".join(["?"] * len(insert_columns))
    insert_sql = f"INSERT INTO imported_raw ({', '.join(insert_columns)}) VALUES ({placeholders})"

    conn = connect()
    inserted = 0
    w_inserted = 0
    with conn:
      conn.execute(
        """DELETE FROM approval_decision WHERE cluster_id IN (
                            SELECT id FROM cluster_proposal WHERE batch_id=?
                        )""",
        (batch_id,),
      )
      conn.execute(
        "DELETE FROM cluster_member WHERE cluster_id IN (SELECT id FROM cluster_proposal WHERE batch_id=?)",
        (batch_id,),
      )
      conn.execute("DELETE FROM cluster_proposal WHERE batch_id=?", (batch_id,))
      conn.execute("DELETE FROM working_article WHERE batch_id=?", (batch_id,))
      conn.execute("DELETE FROM imported_raw WHERE batch_id=?", (batch_id,))

      for excel_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
        row_values = [row[i].value if i < len(row) else None for i in range(len(column_keys))]
        if _row_empty(row_values):
          continue
        record = {column_keys[i]: row_values[i] for i in range(len(column_keys))}
        nome_raw = record.get("nome")
        if nome_raw is None or str(nome_raw).strip() == "":
          continue
        nome = str(nome_raw).strip()
        record["nome"] = nome
        nome_norm = normalize_name(nome)
        now = now_utc()
        features = compute_article_features(nome)

        field_values = {key: record.get(key) for key in CARDEX_FIELD_ORDER}

        values: List[Any] = []
        for col in insert_columns:
          if col == "batch_id":
            values.append(batch_id)
          elif col == "row_index":
            values.append(excel_idx)
          elif col == "nome_norm":
            values.append(nome_norm)
          elif col == "nome_sem_stop":
            values.append(features.nome_sem_stop)
          elif col == "quantidade_valor":
            values.append(features.quantidade_valor)
          elif col == "quantidade_total":
            values.append(features.quantidade_total)
          elif col == "quantidade_unidade":
            values.append(features.quantidade_unidade)
          elif col == "quantidade_tipo":
            values.append(features.quantidade_tipo)
          elif col == "quantidade_numero":
            values.append(features.quantidade_numero)
          elif col == "flag_com_sal":
            values.append(1 if features.flag_com_sal else 0)
          elif col == "flag_sem_sal":
            values.append(1 if features.flag_sem_sal else 0)
          elif col == "marca_detectada":
            values.append(features.marca_detectada)
          elif col in ("desc1", "desc2"):
            values.append(nome)
          elif col == "created_at":
            values.append(now)
          else:
            values.append(field_values.get(col))

        cur = conn.execute(insert_sql, values)
        raw_id = cur.lastrowid
        inserted += 1
        working_values = [
          batch_id,
          raw_id,
          nome_norm,
          features.nome_sem_stop,
          features.quantidade_valor,
          features.quantidade_total,
          features.quantidade_unidade,
          features.quantidade_tipo,
          features.quantidade_numero,
          1 if features.flag_com_sal else 0,
          1 if features.flag_sem_sal else 0,
          features.marca_detectada,
        ]
        conn.execute(
          """INSERT INTO working_article(batch_id, raw_id, nome_norm, nome_sem_stop, quantidade_valor,
                     quantidade_total, quantidade_unidade, quantidade_tipo, quantidade_numero,
                     flag_com_sal, flag_sem_sal, marca_detectada)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
          working_values,
        )
        w_inserted += 1

  finally:
    wb.close()

  return {
    "ok": True,
    "batch_id": batch_id,
    "imported_rows": inserted,
    "working_rows": w_inserted,
    "file_used": xlsx_path,
  }
