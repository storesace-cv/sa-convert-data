from __future__ import annotations

import json
import os
import sqlite3
from typing import Any, Iterable

from openpyxl import load_workbook

from app.backend.audit import current_user, log_action
from app.backend.cardex_schema import CARDEX_FIELD_ORDER
from app.backend.classification_rules import canonical_attrs
from app.backend.db import connect, now_utc
from app.config import EXPORT_DIR, ensure_dirs


CLASSIFICATION_KEYS: tuple[str, ...] = (
    "class_compra",
    "class_venda",
    "class_compra_venda",
    "class_generico",
    "class_producao",
    "class_producao_venda",
)


def _classification_from_canonical(record: dict[str, Any]) -> str:
    canonical_id = record.get("canonical_id")
    if canonical_id:
        scope = record.get("canonical_scope") or "global"
        attrs = canonical_attrs(scope, canonical_id)
        if attrs:
            tag = attrs.get("class_tag")
            if tag == "COMPRA":
                return "class_compra"
            if tag == "COMPRA/VENDA":
                return "class_compra_venda"
        return "class_compra_venda"
    return "class_compra_venda"


def _classification_marks(selected: str) -> dict[str, str | None]:
    return {key: ("X" if key == selected else None) for key in CLASSIFICATION_KEYS}


def _fetch_import_log(conn: sqlite3.Connection, batch_id: str) -> tuple[dict[str, Any], str | None]:
    cur = conn.execute(
        "SELECT payload_json, ts FROM decision_log WHERE action=? ORDER BY id DESC",
        ("cardex_import",),
    )
    for row in cur.fetchall():
        payload = json.loads(row[0] or "{}")
        if payload.get("batch_id") == batch_id:
            return payload, row[1]
    return {}, None


def _count(conn: sqlite3.Connection, sql: str, params: Iterable[Any]) -> int:
    cur = conn.execute(sql, tuple(params))
    row = cur.fetchone()
    return int(row[0] if row and row[0] is not None else 0)


def export_excel_using_model(model_path: str, out_path: str | None, batch_id: str) -> dict[str, Any]:
    ensure_dirs()
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"Modelo não encontrado: {model_path}")
    if not out_path or out_path.strip() == "":
        out_path = str(EXPORT_DIR / f"export-{batch_id}.xlsx")

    wb = load_workbook(model_path)
    ws = wb.active
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

    conn = connect()
    conn.row_factory = sqlite3.Row
    try:
        field_sql = ", ".join(f"ir.{col}" for col in CARDEX_FIELD_ORDER)
        cur = conn.execute(
            f"""
            SELECT
                a.id AS approval_id,
                a.canonical_id,
                COALESCE(ci.name_canonico, ir.nome || ' (m)') AS canonical_label,
                ci.scope AS canonical_scope,
                a.unit_default AS unit_default_final,
                a.unit_compra AS unit_compra_final,
                a.unit_stock AS unit_stock_final,
                a.unit_log AS unit_log_final,
                cp.label_sugerido,
                {field_sql}
            FROM approval_decision a
            JOIN cluster_proposal cp ON cp.id=a.cluster_id
            LEFT JOIN canonical_item ci ON ci.id=a.canonical_id
            LEFT JOIN imported_raw ir ON ir.id=a.artigo_base_raw_id
            WHERE cp.batch_id=?
            ORDER BY cp.id, a.id
            """,
            (batch_id,),
        )
        approvals = [dict(row) for row in cur.fetchall()]

        clusters_total = _count(
            conn,
            "SELECT COUNT(*) FROM cluster_proposal WHERE batch_id=?",
            (batch_id,),
        )
        approvals_total = len(approvals)
        excluded_items = _count(
            conn,
            """
            SELECT COUNT(*)
            FROM cluster_member cm
            JOIN cluster_proposal cp ON cp.id=cm.cluster_id
            WHERE cp.batch_id=? AND IFNULL(cm.selected_by_user, 0)=0
            """,
            (batch_id,),
        )

        import_payload, import_ts = _fetch_import_log(conn, batch_id)
        operator = current_user()
        export_ts = now_utc()
        with conn:
            export_log_id = log_action(
                conn,
                f"batch:{batch_id}",
                "export_generated",
                {
                    "batch_id": batch_id,
                    "out_file": out_path,
                    "rows": approvals_total,
                    "user": operator,
                },
                ts=export_ts,
            )

        row_index = 3
        for record in approvals:
            row_values: dict[str, Any] = {key: record.get(key) for key in CARDEX_FIELD_ORDER}
            canonical_label = record.get("canonical_label") or ""
            for key in ("cod_artigo", "nome", "desc_curta1", "desc_curta2"):
                if key == "cod_artigo":
                    row_values[key] = None
                else:
                    row_values[key] = canonical_label

            class_key = _classification_from_canonical(record)
            row_values.update(_classification_marks(class_key))

            for unit_key, final_key in (
                ("unid_default", "unit_default_final"),
                ("unid_compra", "unit_compra_final"),
                ("unid_stock", "unit_stock_final"),
                ("unid_log", "unit_log_final"),
            ):
                value = record.get(final_key) or row_values.get(unit_key)
                row_values[unit_key] = value

            if row_values.get("categoria") in (None, ""):
                row_values["categoria"] = "Default"
            if row_values.get("estado") in (None, ""):
                row_values["estado"] = "ATIVO"

            for col_index, key in enumerate(CARDEX_FIELD_ORDER, start=1):
                ws.cell(row=row_index, column=col_index).value = row_values.get(key)
            row_index += 1

        if "_LOG" in wb.sheetnames:
            wb.remove(wb["_LOG"])
        log_ws = wb.create_sheet("_LOG")
        log_ws.append(["Chave", "Valor"])

        log_entries = [
            ("batch_id", batch_id),
            ("source_file", import_payload.get("file") or "desconhecido"),
            ("import_timestamp", import_ts or ""),
            ("import_user", import_payload.get("user") or ""),
            ("export_timestamp", export_ts),
            ("export_user", operator),
            ("export_log_id", export_log_id),
            ("export_file", out_path),
            ("clusters_suggested", clusters_total),
            ("clusters_approved", approvals_total),
            ("items_excluded", excluded_items),
        ]

        for key, value in log_entries:
            log_ws.append([key, value])

        wb.save(out_path)
    finally:
        conn.close()

    return {"ok": True, "rows": approvals_total, "out": out_path}
