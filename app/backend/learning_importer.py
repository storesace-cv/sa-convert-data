from __future__ import annotations

import os
import re
import sqlite3
from itertools import combinations
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook

from app.backend.audit import log_action
from app.backend.db import connect, init_db, now_utc
from app.backend.domain_rules import label_from_rules
from app.backend.text_norm import normalize_name
from app.backend.cardex_schema import normalize_header_label


def _resolve_canonical_label(
    scope: str,
    familia: str | None,
    subfamilia: str | None,
    term_norm: str,
) -> str:
    label = label_from_rules(familia, subfamilia, scope=scope)
    if label:
        return label
    head = (term_norm.split(" ") or ["ITEM"])[0]
    return f"{head} (m)"


def _split_synonyms(raw: Any) -> List[str]:
    if raw is None:
        return []
    if isinstance(raw, (list, tuple, set)):
        values = [str(item) for item in raw if str(item).strip()]
    else:
        text = str(raw)
        values = [part.strip() for part in re.split(r"[;,\n]+", text) if part.strip()]
    return values


def _cell_text(row, idx: int | None) -> str:
    if idx is None:
        return ""
    cell = row[idx]
    if cell is None:
        return ""
    value = cell.value
    if value is None:
        return ""
    return str(value).strip()


def _collect_units(row, columns: dict[str, int | None]) -> dict[str, str | None]:
    units: dict[str, str | None] = {}
    for key, idx in columns.items():
        text = _cell_text(row, idx)
        units[key] = text or None
    return units


def _fingerprint(term_norm: str, explicit: str | None) -> str:
    if explicit and explicit.strip():
        return normalize_name(explicit)
    return term_norm


def _normalize_optional(value: str | None) -> str | None:
    if value is None:
        return None
    norm = normalize_name(value)
    return norm or None


def _fetch_canonical_label(conn: sqlite3.Connection, canonical_id: int | None) -> str | None:
    if canonical_id is None:
        return None
    cur = conn.execute(
        "SELECT name_canonico FROM canonical_item WHERE id=?",
        (canonical_id,),
    )
    row = cur.fetchone()
    return row[0] if row else None


def _record_pair_judgements(
    conn: sqlite3.Connection,
    scope: str,
    synonyms: Iterable[str],
) -> int:
    values = sorted({syn for syn in synonyms if syn})
    if len(values) < 2:
        return 0
    updates = 0
    for left, right in combinations(values, 2):
        if left == right:
            continue
        l_key, r_key = (left, right) if left <= right else (right, left)
        try:
            conn.execute(
                """
                INSERT INTO pair_judgement(scope, left_term_norm, right_term_norm, label, support_count, source)
                VALUES(?,?,?,?,?,?)
                """,
                (scope, l_key, r_key, "duplicate", 1, "learning_file"),
            )
        except sqlite3.IntegrityError:
            conn.execute(
                """
                UPDATE pair_judgement
                SET support_count=support_count+1, source='learning_file'
                WHERE scope=? AND left_term_norm=? AND right_term_norm=? AND label='duplicate'
                """,
                (scope, l_key, r_key),
            )
        updates += 1
    return updates


def learn_from_xlsx(xlsx_path: str, scope: str = "global") -> Dict[str, Any]:
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Ficheiro não encontrado: {xlsx_path}")

    init_db()

    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active

        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        headers: List[str] = [str(c.value or "") for c in header_cells]
        normalized_headers = {
            normalize_header_label(header): idx for idx, header in enumerate(headers)
        }

        def require_column(*candidates: str) -> int:
            for candidate in candidates:
                norm = normalize_header_label(candidate)
                if norm in normalized_headers:
                    return normalized_headers[norm]
            raise KeyError(f"Coluna obrigatória não encontrada: {candidates[0]}")

        def optional_column(*candidates: str) -> int | None:
            for candidate in candidates:
                norm = normalize_header_label(candidate)
                if norm in normalized_headers:
                    return normalized_headers[norm]
            return None

        col_nome = require_column("nome")
        col_fam = optional_column("familia")
        col_sub = optional_column("subfamilia")
        col_synonyms = optional_column("sinonimos", "synonyms")
        col_fingerprint = optional_column("fingerprint", "padrao", "pattern")

        unit_columns = {
            "unit_default": optional_column("unidade default", "unit default", "unid default"),
            "unit_compra": optional_column("unidade compra", "unit compra", "unid compra"),
            "unit_stock": optional_column("unidade stock", "unit stock", "unid stock"),
            "unit_log": optional_column(
                "unidade logistica",
                "unit logistica",
                "unit log",
                "unid logistica",
                "unidade log",
            ),
        }

        col_base_hint = optional_column("artigo base", "base article", "base hint")

        conn = connect()

        synonyms_upserted = 0
        canonical_created = 0
        class_map_updates = 0
        group_pattern_updates = 0
        pair_judgement_updates = 0
        logs_recorded = 0

        with conn:
            for row in ws.iter_rows(min_row=2):
                nome_raw = row[col_nome].value if col_nome < len(row) else None
                if nome_raw is None:
                    continue
                nome = str(nome_raw).strip()
                if not nome:
                    continue

                term_norm = normalize_name(nome)
                familia = _normalize_optional(_cell_text(row, col_fam))
                subfamilia = _normalize_optional(_cell_text(row, col_sub))

                existing = conn.execute(
                    "SELECT canonical_id FROM canonical_synonym WHERE scope=? AND term_norm=?",
                    (scope, term_norm),
                ).fetchone()

                canonical_id: int | None = None
                canonical_label: str | None = None

                if existing and existing[0] is not None:
                    canonical_id = existing[0]
                    canonical_label = _fetch_canonical_label(conn, canonical_id)

                if canonical_id is None or canonical_label is None:
                    canonical_label = _resolve_canonical_label(
                        scope, familia, subfamilia, term_norm
                    )
                    cur = conn.execute(
                        "SELECT id FROM canonical_item WHERE name_canonico=? AND scope=?",
                        (canonical_label, scope),
                    )
                    row_item = cur.fetchone()
                    if row_item:
                        canonical_id = row_item[0]
                    else:
                        cur = conn.execute(
                            """
                            INSERT INTO canonical_item(name_canonico, scope, rule_version, created_at)
                            VALUES(?,?,?,?)
                            """,
                            (canonical_label, scope, "v1", now_utc()),
                        )
                        canonical_id = cur.lastrowid
                        canonical_created += 1

                if canonical_label is None:
                    canonical_label = _fetch_canonical_label(conn, canonical_id) or ""

                synonyms = [term_norm]
                if col_synonyms is not None and col_synonyms < len(row):
                    synonyms.extend(_split_synonyms(row[col_synonyms].value))

                unique_synonyms = {normalize_name(value) for value in synonyms if value}
                for synonym in unique_synonyms:
                    if not synonym:
                        continue
                    try:
                        conn.execute(
                            """
                            INSERT INTO canonical_synonym(scope, term_norm, canonical_id, confidence, source)
                            VALUES(?,?,?,?,?)
                            """,
                            (scope, synonym, canonical_id, 1.0, "learning_file"),
                        )
                        synonyms_upserted += 1
                    except sqlite3.IntegrityError:
                        conn.execute(
                            """
                            UPDATE canonical_synonym
                            SET canonical_id=?, confidence=1.0, source='learning_file'
                            WHERE scope=? AND term_norm=?
                            """,
                            (canonical_id, scope, synonym),
                        )
                        synonyms_upserted += 1

                pair_judgement_updates += _record_pair_judgements(
                    conn, scope, unique_synonyms
                )

                if familia or subfamilia:
                    fam_value = familia or ""
                    sub_value = subfamilia or ""
                    try:
                        conn.execute(
                            """
                            INSERT INTO class_map(scope, familia, subfamilia, canonical_label, support_count, source)
                            VALUES(?,?,?,?,?,?)
                            """,
                            (scope, fam_value, sub_value, canonical_label, 1, "learning_file"),
                        )
                        class_map_updates += 1
                    except sqlite3.IntegrityError:
                        conn.execute(
                            """
                            UPDATE class_map
                            SET canonical_label=?, support_count=support_count+1
                            WHERE scope=? AND familia=? AND subfamilia=?
                            """,
                            (canonical_label, scope, fam_value, sub_value),
                        )
                        class_map_updates += 1

                fingerprint_value = _fingerprint(
                    term_norm,
                    _cell_text(row, col_fingerprint),
                )

                units = _collect_units(row, unit_columns)
                base_hint = _cell_text(row, col_base_hint) or nome

                try:
                    conn.execute(
                        """
                        INSERT INTO group_pattern(
                            scope, fingerprint, canonical_id,
                            unit_default, unit_compra, unit_stock, unit_log,
                            base_article_hint, support_count, source
                        ) VALUES(?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            scope,
                            fingerprint_value,
                            canonical_id,
                            units.get("unit_default"),
                            units.get("unit_compra"),
                            units.get("unit_stock"),
                            units.get("unit_log"),
                            base_hint,
                            1,
                            "learning_file",
                        ),
                    )
                    group_pattern_updates += 1
                except sqlite3.IntegrityError:
                    conn.execute(
                        """
                        UPDATE group_pattern
                        SET canonical_id=?,
                            unit_default=COALESCE(?, unit_default),
                            unit_compra=COALESCE(?, unit_compra),
                            unit_stock=COALESCE(?, unit_stock),
                            unit_log=COALESCE(?, unit_log),
                            base_article_hint=COALESCE(?, base_article_hint),
                            support_count=support_count+1
                        WHERE scope=? AND fingerprint=?
                        """,
                        (
                            canonical_id,
                            units.get("unit_default"),
                            units.get("unit_compra"),
                            units.get("unit_stock"),
                            units.get("unit_log"),
                            base_hint,
                            scope,
                            fingerprint_value,
                        ),
                    )
                    group_pattern_updates += 1

                log_action(
                    conn,
                    scope,
                    "learning_upsert",
                    {
                        "term_norm": term_norm,
                        "canonical_id": canonical_id,
                        "canonical_label": canonical_label,
                        "synonyms": sorted(unique_synonyms),
                        "familia": familia,
                        "subfamilia": subfamilia,
                        "fingerprint": fingerprint_value,
                    },
                )
                logs_recorded += 1
    finally:
        wb.close()

    return {
        "ok": True,
        "file": xlsx_path,
        "scope": scope,
        "synonyms_upserted": synonyms_upserted,
        "canonical_created": canonical_created,
        "class_map_updates": class_map_updates,
        "group_pattern_updates": group_pattern_updates,
        "pair_judgement_updates": pair_judgement_updates,
        "logs_recorded": logs_recorded,
    }


def forget_learning(scope: str = "global") -> Dict[str, Any]:
    init_db()
    conn = connect()

    with conn:
        def _deleted(cur: sqlite3.Cursor) -> int:
            return cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0

        synonyms_deleted = _deleted(
            conn.execute(
                "DELETE FROM canonical_synonym WHERE scope=? AND source='learning_file'",
                (scope,),
            )
        )

        class_map_deleted = _deleted(
            conn.execute(
                "DELETE FROM class_map WHERE scope=? AND source='learning_file'",
                (scope,),
            )
        )

        group_patterns_deleted = _deleted(
            conn.execute(
                "DELETE FROM group_pattern WHERE scope=? AND source='learning_file'",
                (scope,),
            )
        )

        pair_judgements_deleted = _deleted(
            conn.execute(
                "DELETE FROM pair_judgement WHERE scope=? AND source='learning_file'",
                (scope,),
            )
        )

        logs_deleted = _deleted(
            conn.execute(
                "DELETE FROM decision_log WHERE scope=? AND action='learning_upsert'",
                (scope,),
            )
        )

        canonical_items_deleted = _deleted(
            conn.execute(
                """
                DELETE FROM canonical_item
                WHERE scope=?
                  AND NOT EXISTS (
                      SELECT 1 FROM canonical_synonym cs WHERE cs.canonical_id=canonical_item.id
                  )
                  AND NOT EXISTS (
                      SELECT 1 FROM group_pattern gp WHERE gp.canonical_id=canonical_item.id
                  )
                  AND NOT EXISTS (
                      SELECT 1 FROM approval_decision ad WHERE ad.canonical_id=canonical_item.id
                  )
                """,
                (scope,),
            )
        )

        payload = {
            "synonyms_deleted": synonyms_deleted,
            "class_map_deleted": class_map_deleted,
            "group_patterns_deleted": group_patterns_deleted,
            "canonical_items_deleted": canonical_items_deleted,
            "logs_deleted": logs_deleted,
            "pair_judgements_deleted": pair_judgements_deleted,
        }

        action_id = log_action(conn, scope, "learning_forget", payload)

    return {
        "ok": True,
        "scope": scope,
        **payload,
        "action_id": action_id,
    }
